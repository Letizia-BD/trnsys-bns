import os
from openpyxl import load_workbook
os.environ["JULIA_BINDIR"] = r"C:\Users\ITM User\AppData\Local\Programs\Julia-1.11.5\bin"
from juliacall import Main as jl
sys.path.insert(1, "C:/Users/ITM User/.julia/packages/BoreholeNetworksSimulator/G7Ojx")
import BNSPythonAdapter.src.adapter
import sys
import json
import numpy as np

import sys
import pickle
from juliacall import Main as jl
import numpy as np



# # Load borehole properties from Excel
# wb = load_workbook("GeoInput.xlsx")
# sheet = wb['Borehole']

# filled_rows = 0
# for row in sheet.iter_rows():
#     if any(cell.value is not None for cell in row):
#         filled_rows += 1

# Nb = filled_rows - 1
# H_list = [sheet[f"A{row}"].value for row in range(2, filled_rows + 1)]
# D_list = [sheet[f"B{row}"].value for row in range(2, filled_rows + 1)]
# x_list = [sheet[f"D{row}"].value for row in range(2, filled_rows + 1)]
# y_list = [sheet[f"E{row}"].value for row in range(2, filled_rows + 1)]
# activation_year_list = [sheet[f"F{row}"].value for row in range(2, filled_rows + 1)]

# unique_years = sorted(set(activation_year_list))
# if len(unique_years) != 2:
#     sys.exit("Error: Exactly 2 unique activation years required.")

# Nb1 = activation_year_list.count(unique_years[0])

# # Load ground properties
# sheet = wb['Ground']
# λ = float(sheet['A2'].value)
# rho = float(sheet['B2'].value)
# cp = float(sheet['C2'].value)
# alpha = λ / (rho * cp)
# T0 = float(sheet['E2'].value)

# # Simulation parameters
# # dt = TRNData[thisModule]["simulation time step"] * 3600.
# # Nt = TRNData[thisModule]["total number of time steps"]
# dt = 3600.
# Nt = 43800
# activation_step = int(unique_years[1] * 8760 * 3600. / dt)

# # Fluid
# fluid = jl.Water()

# # Borehole positions
# positions = jl.Array[jl.Tuple[jl.Float64, jl.Float64]]([(float(x), float(y)) for x, y in zip(x_list, y_list)])

# # Mass flows
# mass_flows = jl.Array[jl.Float64](0.6 * np.ones(Nb))

# # Networks
# network_1 = jl.BoreholeNetwork(Nb)
# n1_count = 0
# for i in range(Nb):
#     if activation_year_list[i] == activation_year_list[0]:
#         jl.connect_to_source_b(network_1, i+1)
#         jl.connect_to_sink_b(network_1, i+1)
#         n1_count += 1
#     else:
#         jl.connect_b(network_1, i+1, i+1)

# network_2 = jl.BoreholeNetwork(Nb)
# for i in range(Nb):
#     jl.connect_to_source_b(network_2, i+1)
#     jl.connect_to_source_b(network_2, i+1)
#     jl.connect_to_sink_b(network_2, i+1)
#     jl.connect_to_sink_b(network_2, i+1)

# configurations = jl.Vector([network_1, network_2])

# method = jl.NonHistoryMethod()
# medium = jl.GroundMedium(λ=λ, α=alpha, T0=T0)

# py_boreholes = [jl.SingleUPipeBorehole(H=float(H_list[i]), D=float(D_list[i])) for i in range(Nb)]
# boreholes = jl.Array(py_boreholes)
# borefield = jl.HeterogeneousBorefield(boreholes=boreholes, positions=positions)

# T_initial = jl.Array[jl.Float64]([-3.0 for _ in range(1, Nt+1)])
# constraint = jl.uniform_InletTempConstraint(T_initial, Nb)

# options = jl.SimulationOptions(
#     method=method,
#     constraint=constraint,
#     borefield=borefield,
#     fluid=fluid,
#     medium=medium,
#     boundary_condition=jl.DirichletBoundaryCondition(),
#     Δt=dt,
#     Nt=Nt,
#     configurations=configurations
# )

# containers = jl.initialize(options)

# class StepOperator():
#     def __init__(self, mass_flows, activation_step,Nb):
#         self.mass_flow_containers = jl.Array[jl.Float64](np.zeros(Nb))
#         self.mass_flows = mass_flows
#         self.activation_step = activation_step
#         self.Tin = 0

#     def update(self, Tin, mf):
#         self.Tin = Tin
#         self.mass_flows[:] = mf

#     def operate(self,step, options,X):
#         options.constraint.T_in[:, step] = self.Tin
#         after_step = step >= self.activation_step
#         active_configuration = 1 if after_step else 0
#         active_network = options.configurations[active_configuration]

#         options.constraint.T_in[:, step] = self.Tin

#         if after_step:
#             # self.mass_flow_containers[:] = self.mass_flows/Nb  
#             for i in range(Nb):
#                     self.mass_flow_containers[i] = self.mass_flows[i]/Nb
#         else:
#             for i in range(Nb):
#                 if i in range(N1):
#                     self.mass_flow_containers[i] = self.mass_flows[i]/Nb1
#                 else:
#                     self.mass_flow_containers[i] = 0.0

#         return jl.BoreholeOperation(network=active_network, mass_flows=self.mass_flow_containers)

# # Operator class already in Julia
# operator = StepOperator(mass_flows, activation_step, Nb)
# # Run a single Julia step
# # Read inputs from command-line
# Tin = float(sys.argv[1])
# m = float(sys.argv[2])
# stepNo = int(sys.argv[3])

# operator.update(Tin, m)

# jl.simulate_steps_b(n=1, initial_step=stepNo, operator=operator, options=options, containers=containers)

# # Compute output
# if stepNo < activation_step:
#     Tout = np.mean(containers.X[1:Nb1*2+1:2, stepNo-1])
# else:
#     Tout = np.mean(containers.X[1:Nb*2+1:2, stepNo-1])

Tout = Tin
# Return the result to TRNSYS
print(Tout)
sys.stdout.flush()

