# Python module for the TRNSYS Type calling Python using CFFI
# Data exchange with TRNSYS uses a dictionary, called TRNData in this file (it is the argument of all functions).
# Data for this module will be in a nested dictionary under the module name,
# i.e. if this file is calle "MyScript.py", the inputs will be in TRNData["MyScript"]["inputs"]
# for convenience the module name is saved in thisModule
#
# MKu, 2022-02-15

import os
import pathlib
import sys
import io
import contextlib
import numpy as np
import pandas as pd

from openpyxl import load_workbook

# Set environment variables for Julia
os.environ["JULIA_BINDIR"] = r"C:\Users\ITM User\AppData\Local\Programs\Julia-1.11.5\bin"
os.environ["JULIA_DEPOT_PATH"] = r"C:\MyJuliaDepot"
os.environ["JULIA_NUM_THREADS"] = "1"

from juliacall import Main as jl
from juliacall import Pkg as jlPkg

# Old lines for finding the package 
# path = jl.Base.find_package("BoreholeNetworksSimulator")
# sys.path.insert(1, "C:/BoreholeNetworksSimulator.jl")

# Initialize Julia (once at DLL startup)
jl.seval("""
using BoreholeNetworksSimulator
""")

sys.path.insert(1, r"C:\MyJuliaDepot\packages\BoreholeNetworksSimulator\G7Ojx")
import BNSPythonAdapter.src.adapter

thisModule = os.path.splitext(os.path.basename(__file__))[0]


# Initialization: function called at TRNSYS initialization
# ---------------------------------------------------------------------------------------------------------------------
def Initialization(TRNData):
    global operator, options, containers
    global activation_step
    global N1, Nb
    global Tout

    # Load the borehole properties
    wb = load_workbook("GeoInput.xlsx")
    sheet = wb['Borehole']

    filled_rows = 0
    for row in sheet.iter_rows():
        if any(cell.value is not None for cell in row):
            filled_rows += 1

    # Number of boreholes
    Nb = filled_rows - 1

    H_list = [None] * (Nb)
    D_list = [None] * (Nb)
    rb_list = [None] * (Nb)
    x_list = [None] * (Nb)
    y_list = [None] * (Nb)
    activation_year_list = [None] * (Nb)

    for row in range(2,filled_rows + 1):
        H_list[row-2] = sheet[f"A{row}"].value
        D_list[row-2] = sheet[f"B{row}"].value  
        rb_list[row-2] = sheet[f"C{row}"].value  
        x_list[row-2] = sheet[f"D{row}"].value
        y_list[row-2] = sheet[f"E{row}"].value   
        activation_year_list[row-2] = sheet[f"F{row}"].value  

    unique_years = set(activation_year_list)
    num_unique_years = len(unique_years) 

    if num_unique_years !=2:
        sys.exit("Error: The number of unique activation years must be exactly 2.")

    Nb1 = activation_year_list.count(min(unique_years))

    # Load the ground properties
    sheet = wb['Ground']
    λ = float(sheet['A2'].value)
    rho = float(sheet['B2'].value)
    cp = float(sheet['C2'].value)

    # Ground thermal diffusivity
    alpha = λ/rho/cp # [m2/s]

    T0 = float(sheet['E2'].value)

    # Simulation parameters (must be consistent with TRNSYS!)

    # Time step in seconds
    dt = TRNData[thisModule]["simulation time step"] * 3600.
    # Number of time steps
    Nt = TRNData[thisModule]["total number of time steps"]
    
    # Simulation step at which the second netork starts
    activation_step = int(list(unique_years)[1] * 8760 * 3600./dt)

    # Inlet temperature to the borehole(s) to initialize the model. Can be overwritten at each time step.
    Tin = -3. # [degC]
    m = 1.2

    sheet = wb['Fluid']
    fluid_name= sheet['A2'].value
    fluid = jl.seval(f"{fluid_name}()")

    positions = jl.Array[jl.Tuple[jl.Float64, jl.Float64]]([
    (float(x), float(y)) for x, y in zip(x_list, y_list)
    ])

    # total_mass_flow = 1. #[kg/s] or [l/s]?
    mass_flows = jl.Array[jl.Float64](0.6 * np.ones(Nb))
    
    ### Initialize problem - no need for user intervention
    # In network_1 only borehole some boreholes operates, and the remaining boreholes does not exist/operate
    network_1 = jl.BoreholeNetwork(Nb)
    N1 = 0
    for i in range(Nb):
        if activation_year_list[i] == activation_year_list[0]:
            jl.connect_to_source_b(network_1, i+1)
            jl.connect_to_sink_b(network_1, i+1)
            N1+=1
        else:
            jl.connect_b(network_1, i+1, i+1)

    # In network_2 all the boreholes operate in parallel
    network_2 = jl.BoreholeNetwork(Nb)
    for i in range(Nb):
        jl.connect_to_source_b(network_2, i+1)
        jl.connect_to_source_b(network_2, i+1)
        jl.connect_to_sink_b(network_2, i+1)
        jl.connect_to_sink_b(network_2, i+1)

    configurations = jl.Vector([network_1, network_2])

    if max(H_list) > 150:
        method = jl.ConvolutionMethod()
    else:
        method = jl.OriginalNonHistoryMethod()

    # The following method is the fastest and most accurate but makes the TRNSYS simulation crash 
    # method = jl.NonHistoryMethod()

    medium = jl.GroundMedium(λ=λ, α=alpha, T0=T0)

    py_boreholes = []

    for i in range(Nb):
        borehole = jl.SingleUPipeBorehole(H = float(H_list[i]), D = float(D_list[i]))
        py_boreholes.append(borehole)

    boreholes = jl.Array(py_boreholes)

    # Create the borefield object 
    borefield = jl.HeterogeneousBorefield(boreholes=boreholes, positions=positions)

    # Define the boundary condition
    T_initial = jl.Array[jl.Float64]([Tin for i in range(1, Nt+1)])
    constraint = jl.uniform_InletTempConstraint(T_initial, Nb)

    options = jl.SimulationOptions(
        method = method,
        constraint = constraint,
        borefield = borefield,
        fluid = fluid,
        medium = medium,
        boundary_condition = jl.DirichletBoundaryCondition(),
        Δt = dt,
        Nt = Nt,
        configurations = configurations
    )

    containers = jl.initialize(options)

    class StepOperator():
        def __init__(self, mass_flows, activation_step,Nb):
            self.mass_flow_containers = jl.Array[jl.Float64](np.zeros(Nb))
            self.mass_flows = mass_flows
            self.activation_step = activation_step
            self.Tin = 0

        def update(self, Tin, mf):
            self.Tin = Tin
            self.mass_flows[:] = mf

        def operate(self,step, options,X):
            options.constraint.T_in[:, step] = self.Tin
            after_step = step >= self.activation_step
            active_configuration = 1 if after_step else 0
            active_network = options.configurations[active_configuration]

            options.constraint.T_in[:, step] = self.Tin

            if after_step:
                for i in range(Nb):
                        self.mass_flow_containers[i] = self.mass_flows[i]/Nb
            else:
                for i in range(Nb):
                    if i in range(N1):
                        self.mass_flow_containers[i] = self.mass_flows[i]/Nb1
                    else:
                        self.mass_flow_containers[i] = 0.0

            return jl.BoreholeOperation(network=active_network, mass_flows=self.mass_flow_containers)

    operator = StepOperator(mass_flows, activation_step,Nb)

    return

# StartTime: function called at TRNSYS starting time (not an actual time step, initial values should be reported)
# ----------------------------------------------------------------------------------------------------------------------
def StartTime(TRNData):
    return

# Iteration: function called at each TRNSYS iteration within a time step
# ----------------------------------------------------------------------------------------------------------------------
def Iteration(TRNData):

    stepNo = TRNData[thisModule]["current time step number"]
    Tin = TRNData[thisModule]["inputs"][0]
    m = TRNData[thisModule]["inputs"][1]
    
    operator.update(Tin, m)

    jl.simulate_steps_b(n = 1, initial_step = stepNo, operator=operator, options=options, containers=containers)
    if stepNo < activation_step:
        Tout = containers.X[1:N1*2+1:2, stepNo-1]  
        # Tin = containers.X[0:N1*2:2, stepNo-1]  
        # Tb = containers.X[2*Nb:2*Nb+N1:1,stepNo-1]

    else:
        Tout = containers.X[1:Nb*2+1:2, stepNo-1]
        # Tin = containers.X[0:Nb*2:2, stepNo-1]  
        # Tb = containers.X[2*Nb:3*Nb:1,stepNo-1]

    # # Set outputs in TRNData
    TRNData[thisModule]["outputs"][0] = np.mean(Tout)

    with open("Result.txt","a") as file:
        file.write(str(np.mean(Tout))+"\n")
    # --- Outlet temperature assuming equally distributed flow among the boreholes ---
    
    # TRNData[thisModule]["outputs"][1] = Q_tot[stepNo -1]

    return


# EndOfTimeStep: function called at the end of each time step, after iteration and before moving on to next time step
# ----------------------------------------------------------------------------------------------------------------------
def EndOfTimeStep(TRNData):

    # This model has nothing to do during the end-of-step call
    
    return


# LastCallOfSimulation: function called at the end of the simulation (once) - outputs are meaningless at this call
# ----------------------------------------------------------------------------------------------------------------------
def LastCallOfSimulation(TRNData):

    # with open("Q1.txt","a") as file:
    #     print(containers.X[3 * Nb, :], file=file)
    # with open("Q2.txt","a") as file:
    #     print(containers.X[3 * Nb+1, :], file=file)
    # with open("Q3.txt","a") as file:
    #     print(containers.X[3 * Nb+2, :], file=file)
    # with open("Q4.txt","a") as file:
    #     print(containers.X[3 * Nb+3, :], file=file)
    with open("Tf_out.txt","a") as file:
        print(Tout, file=file)


    # NOTE: TRNSYS performs this call AFTER the executable (the online plotter if there is one) is closed. 
    # Python errors in this function will be difficult (or impossible) to diagnose as they will produce no message.
    # A recommended alternative for "end of simulation" actions it to implement them in the EndOfTimeStep() part, 
    # within a condition that the last time step has been reached.
    #
    # Example (to be placed in EndOfTimeStep()):
    #
    # stepNo = TRNData[thisModule]["current time step number"]
    # nSteps = TRNData[thisModule]["total number of time steps"]
    # if stepNo == nSteps-1:     # Remember: TRNSYS steps go from 0 to (number of steps - 1)
    #     do stuff that needs to be done only at the end of simulation

    return