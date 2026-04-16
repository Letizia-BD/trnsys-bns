import pytest
from juliacall import Main as jl
from juliacall import Pkg as jlPkg
import pathlib
from pathlib import Path
import sys, os
import io
import contextlib
import numpy as np
import pandas as pd
from openpyxl import load_workbook

jl.seval("""
using BoreholeNetworksSimulatorFork
""")

path = Path(jl.Base.find_package("BoreholeNetworksSimulatorFork"))
pkg_root = path.parents[1] 
sys.path.insert(0, str(pkg_root))
import BNSPythonAdapter.src.adapter

def python_results():
    # Load the borehole properties
    wb = load_workbook("GeoInput.xlsx")
    sheet = wb['Borehole']

    filled_rows = 0
    for row in sheet.iter_rows():
        if any(cell.value is not None for cell in row):
            filled_rows += 1

    # Number of boreholes
    Nb = filled_rows - 1

    H_list = [None] * (Nb)
    D_list = [None] * (Nb)
    rb_list = [None] * (Nb)
    x_list = [None] * (Nb)
    y_list = [None] * (Nb)
    activation_year_list = [None] * (Nb)

    for row in range(2,filled_rows + 1):
        H_list[row-2] = sheet[f"A{row}"].value
        D_list[row-2] = sheet[f"B{row}"].value  
        rb_list[row-2] = sheet[f"C{row}"].value  
        x_list[row-2] = sheet[f"D{row}"].value
        y_list[row-2] = sheet[f"E{row}"].value   
        activation_year_list[row-2] = sheet[f"F{row}"].value  

    unique_years = set(activation_year_list)
    num_unique_years = len(unique_years) 


    if num_unique_years !=2:
        sys.exit("Error: The number of unique activation years must be exactly 2.")

    Nb1 = activation_year_list.count(min(unique_years))

    # Load the ground properties
    sheet = wb['Ground']
    λ = float(sheet['A2'].value)
    rho = float(sheet['B2'].value)
    cp = float(sheet['C2'].value)

    # Ground thermal diffusivity
    alpha = λ/rho/cp # [m2/s]

    T0 = float(sheet['E2'].value)

    # Simulation parameters (must be consistent with TRNSYS!)

    # Time step in seconds
    dt =  3600.
    # Number of time steps (
    Nt = 43800
    # Simulation step at which the second netork starts
    activation_step = int(list(unique_years)[1] * 8760 * 3600./dt)

    # Inlet temperature to the borehole(s) to initialize the model. Can be overwritten at each time step.
    Tin = -3. # [degC]
    m = 1.2

    # Need to check about possible fluids
    # fluid = jl.Water()
    sheet = wb['Fluid']
    fluid_name= sheet['A2'].value
    fluid = jl.seval(f"{fluid_name}()")

    # positions = jl.Array[jl.Tuple[jl.Float64, jl.Float64]]([(0., 0.), (0., σ)])
    positions = jl.Array[jl.Tuple[jl.Float64, jl.Float64]]([
    (float(x), float(y)) for x, y in zip(x_list, y_list)
    ])

    # total_mass_flow = 1. #[kg/s] or [l/s]?
    mass_flows = jl.Array[jl.Float64](0.6 * np.ones(Nb))
    
    ### Initialize problem - no need for user intervention
    # In network_1 only borehole some boreholes operates, and the remaining boreholes does not exist/operate
    network_1 = jl.BoreholeNetwork(Nb)
    N1 = 0
    for i in range(Nb):
        if activation_year_list[i] == activation_year_list[0]:
            jl.connect_to_source_b(network_1, i+1)
            jl.connect_to_sink_b(network_1, i+1)
            N1+=1
        else:
            jl.connect_b(network_1, i+1, i+1)

    # In network_2 all the boreholes operate in parallel
    network_2 = jl.BoreholeNetwork(Nb)
    for i in range(Nb):
        jl.connect_to_source_b(network_2, i+1)
        jl.connect_to_source_b(network_2, i+1)
        jl.connect_to_sink_b(network_2, i+1)
        jl.connect_to_sink_b(network_2, i+1)

    configurations = jl.Vector([network_1, network_2])


    medium = jl.GroundMedium(λ=λ, α=alpha, T0=T0)

    py_boreholes = []

    for i in range(Nb):
        borehole = jl.SingleUPipeBorehole(H = float(H_list[i]), D = float(D_list[i]))
        py_boreholes.append(borehole)

    boreholes = jl.Array(py_boreholes)

    # Create the borefield object 
    borefield = jl.HeterogeneousBorefield(boreholes=boreholes, positions=positions)

    # D = 0.              # Borehole buried depth (m)
    # H = 150.            # Borehole active length (m)

    # Nb = 4              # Number of boreholes. Obs! If you change this value you have to change the positions and networks 
    # σ = 5.                              # Distance between boreholes

    # positions = jl.Array[jl.Tuple[jl.Float64, jl.Float64]]([(0., 0.), (0., σ),(σ, 0.), (σ, σ)])     # Coordinates of the two boreholes
    # # Create the borehole object 
    # borehole = jl.SingleUPipeBorehole(H=H, D=D)

    # # Create the borefield object 
    # borefield = jl.EqualBoreholesBorefield(borehole_prototype=borehole, positions=positions)


    # Define the boundary condition
    T_initial = jl.Array[jl.Float64]([Tin for i in range(1, Nt+1)])
    constraint = jl.uniform_InletTempConstraint(T_initial, Nb)

    if max(H_list) > 150:
        method = jl.ConvolutionMethod()
    else:
        method = jl.OriginalNonHistoryMethod()

    options = jl.SimulationOptions(
        method = method,
        constraint = constraint,
        borefield = borefield,
        fluid = fluid,
        medium = medium,
        boundary_condition = jl.DirichletBoundaryCondition(),
        Δt = dt,
        Nt = Nt,
        configurations = configurations
    )

    containers = jl.initialize(options)

    class StepOperator():
        def __init__(self, mass_flows, activation_step,Nb):
            self.mass_flow_containers = jl.Array[jl.Float64](np.zeros(Nb))
            self.mass_flows = mass_flows
            self.activation_step = activation_step
            self.Tin = 0

        def update(self, Tin, mf):
            self.Tin = Tin
            self.mass_flows[:] = mf

        def operate(self,step, options,X):
            options.constraint.T_in[:, step] = self.Tin
            after_step = step >= self.activation_step
            active_configuration = 1 if after_step else 0
            active_network = options.configurations[active_configuration]

            options.constraint.T_in[:, step] = self.Tin

            if after_step:
                # self.mass_flow_containers[:] = self.mass_flows/Nb  
                for i in range(Nb):
                        self.mass_flow_containers[i] = self.mass_flows[i]/Nb
            else:
                for i in range(Nb):
                    if i in range(N1):
                        self.mass_flow_containers[i] = self.mass_flows[i]/Nb1
                    else:
                        self.mass_flow_containers[i] = 0.0

            return jl.BoreholeOperation(network=active_network, mass_flows=self.mass_flow_containers)


    operator = StepOperator(mass_flows, activation_step,Nb)
    Tout = []

    for i in range(1,Nt+1):
        Tin = -3
        m = 1.2

        operator.update(Tin, m)

        jl.simulate_steps_b(n = 1, initial_step = i, operator=operator, options=options, containers=containers)

        if i < activation_step:
            Tout.append(np.mean(containers.X[1:N1*2+1:2, i-1]))
        else:
            Tout.append(np.mean(containers.X[1:Nb*2+1:2, i-1]))
        
    with open("Tout_python_debug.txt", "w") as f:
        f.write(",".join(map(str, Tout)))

if __name__ == "__main__":
    python_results()